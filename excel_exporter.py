from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, BarChart, Reference
from typing import List, Dict
import os
from datetime import datetime
from sentiment_analyzer import SentimentAnalyzer


class ExcelExporter:
    """Экспорт отзывов в Excel файл"""

    def __init__(self):
        self.output_dir = "exports"
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        self.sentiment_analyzer = SentimentAnalyzer()

    @staticmethod
    def sanitize_cell_value(value: str) -> str:
        """
        Защита от CSV/Excel injection
        Удаляет опасные символы в начале строки
        """
        if not isinstance(value, str):
            return value

        # Опасные символы которые могут начинать формулу
        dangerous_chars = ['=', '+', '-', '@', '\t', '\r']

        if value and value[0] in dangerous_chars:
            # Добавляем пробел в начало чтобы обезвредить формулу
            return ' ' + value

        return value

    def export_reviews(self, reviews: List[Dict], marketplace: str) -> str:
        """
        Экспортирует отзывы в Excel файл с анализом тональности

        Args:
            reviews: список отзывов в формате словаря
            marketplace: название маркетплейса (WB или Ozon)

        Returns:
            путь к созданному файлу
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Отзывы"

        # Проводим анализ тональности
        analysis = self.sentiment_analyzer.analyze_reviews_batch(reviews)

        # Заголовки столбцов (добавили колонку Тональность)
        headers = ['Автор', 'Рейтинг', 'Дата', 'Текст отзыва', 'Плюсы', 'Минусы', 'Тональность']

        # Стилизация заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Записываем данные с защитой от инъекций
        for row_num, review in enumerate(reviews, 2):
            ws.cell(row=row_num, column=1, value=self.sanitize_cell_value(review.get('Автор', '')))
            ws.cell(row=row_num, column=2, value=review.get('Рейтинг', 0))
            ws.cell(row=row_num, column=3, value=review.get('Дата', ''))
            ws.cell(row=row_num, column=4, value=self.sanitize_cell_value(review.get('Текст отзыва', '')))
            ws.cell(row=row_num, column=5, value=self.sanitize_cell_value(review.get('Плюсы', '')))
            ws.cell(row=row_num, column=6, value=self.sanitize_cell_value(review.get('Минусы', '')))

            # Добавляем тональность
            sentiment_result = self.sentiment_analyzer.analyze_review(
                review.get('Текст отзыва', ''),
                review.get('Рейтинг', 0)
            )
            sentiment_text = sentiment_result['sentiment']
            sentiment_emoji = {'Позитивный': '😊', 'Негативный': '😞', 'Нейтральный': '😐'}.get(sentiment_text, '😐')

            sentiment_cell = ws.cell(row=row_num, column=7, value=f"{sentiment_emoji} {sentiment_text}")

            # Цветовое выделение тональности
            if sentiment_text == 'Позитивный':
                sentiment_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif sentiment_text == 'Негативный':
                sentiment_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                sentiment_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        # Настройка ширины столбцов
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 18

        # Выравнивание текста
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Создаем лист с анализом тональности
        self._create_analysis_sheet(wb, analysis)

        # Генерируем имя файла с текущей датой и временем
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{marketplace}_reviews_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        # Сохраняем файл
        wb.save(filepath)

        return filepath

    def _create_analysis_sheet(self, wb: Workbook, analysis: Dict):
        """Создает лист с анализом тональности и графиками"""
        ws = wb.create_sheet(title="Анализ тональности")

        # Стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)

        # Заголовок
        ws['A1'] = "📊 Анализ тональности отзывов"
        ws['A1'].font = title_font
        ws.merge_cells('A1:B1')

        # Общая статистика
        row = 3
        ws[f'A{row}'] = "Общая статистика"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        stats = [
            ('Всего отзывов:', analysis['total_reviews']),
            ('Позитивных:', f"{analysis['positive_reviews']} ({analysis['positive_percentage']:.1f}%)"),
            ('Негативных:', f"{analysis['negative_reviews']} ({analysis['negative_percentage']:.1f}%)"),
            ('Нейтральных:', f"{analysis['neutral_reviews']} ({analysis['neutral_percentage']:.1f}%)"),
            ('Средний рейтинг:', f"{analysis['average_rating']:.2f} ⭐"),
        ]

        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Отзывы с проблемами
        row += 1
        ws[f'A{row}'] = "Проблемные отзывы"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = "Отзывы с упоминанием проблем:"
        ws[f'B{row}'] = analysis['reviews_with_problems']
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        # Ключевые слова
        row += 1
        ws[f'A{row}'] = "🔑 Топ-10 ключевых слов"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = "Слово"
        ws[f'B{row}'] = "Количество упоминаний"
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        row += 1

        for word, count in analysis['top_keywords']:
            ws[f'A{row}'] = word
            ws[f'B{row}'] = count
            row += 1

        # Настройка ширины столбцов
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 35

        # Создаем круговую диаграмму распределения тональности
        self._create_sentiment_pie_chart(ws, analysis)

        # Создаем столбчатую диаграмму распределения рейтингов
        self._create_rating_bar_chart(ws, analysis)

    def _create_sentiment_pie_chart(self, ws, analysis: Dict):
        """Создает круговую диаграмму распределения тональности"""
        # Данные для диаграммы
        chart_row = ws.max_row + 2
        ws[f'D{chart_row}'] = "Тип"
        ws[f'E{chart_row}'] = "Количество"

        ws[f'D{chart_row + 1}'] = "Позитивные"
        ws[f'E{chart_row + 1}'] = analysis['positive_reviews']

        ws[f'D{chart_row + 2}'] = "Негативные"
        ws[f'E{chart_row + 2}'] = analysis['negative_reviews']

        ws[f'D{chart_row + 3}'] = "Нейтральные"
        ws[f'E{chart_row + 3}'] = analysis['neutral_reviews']

        # Создаем диаграмму
        chart = PieChart()
        chart.title = "Распределение тональности отзывов"
        chart.style = 10
        chart.height = 10
        chart.width = 15

        labels = Reference(ws, min_col=4, min_row=chart_row + 1, max_row=chart_row + 3)
        data = Reference(ws, min_col=5, min_row=chart_row, max_row=chart_row + 3)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        ws.add_chart(chart, f"D{chart_row + 5}")

    def _create_rating_bar_chart(self, ws, analysis: Dict):
        """Создает столбчатую диаграмму распределения рейтингов"""
        # Данные для диаграммы
        chart_row = ws.max_row + 18
        ws[f'D{chart_row}'] = "Рейтинг"
        ws[f'E{chart_row}'] = "Количество"

        # Подсчитываем распределение рейтингов
        rating_distribution = analysis.get('rating_distribution', {})

        for rating in range(1, 6):
            ws[f'D{chart_row + rating}'] = f"{rating} ⭐"
            ws[f'E{chart_row + rating}'] = rating_distribution.get(rating, 0)

        # Создаем диаграмму
        chart = BarChart()
        chart.title = "Распределение рейтингов"
        chart.style = 10
        chart.height = 10
        chart.width = 15

        labels = Reference(ws, min_col=4, min_row=chart_row + 1, max_row=chart_row + 5)
        data = Reference(ws, min_col=5, min_row=chart_row, max_row=chart_row + 5)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        chart.x_axis.title = "Рейтинг"
        chart.y_axis.title = "Количество отзывов"

        ws.add_chart(chart, f"D{chart_row + 7}")
